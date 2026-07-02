import re
from pathlib import Path
from datetime import datetime, date, time
from openpyxl import load_workbook


FORM_FILE = "問卷表單.xlsx"
TEMPLATE_FILE = "立人採檢預約單_外縣市1140722新版.xlsx"
OUTPUT_DIR = "產生結果"

SHEET_NAME = "委託單-個別通知"

CELLS = {
    "姓名": "D5",
    "身分證": "D6",
    "生日": "D7",
    "檢驗所": "D9",
    "地址": "D10",
    "年": "C11",
    "月": "E11",
    "日": "G11",
    "時": "C12",
    "分": "F12",
}


def clean(x):
    return "" if x is None else str(x).strip()


def safe_name(x):
    return re.sub(r'[\\/:*?"<>|\s]+', "-", clean(x))


def parse_date(x):
    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, date):
        return x

    x = clean(x)
    for fmt in ["%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d", "%Y%m%d"]:
        try:
            return datetime.strptime(x, fmt).date()
        except ValueError:
            pass

    raise ValueError(f"日期格式錯誤：{x}")


def parse_time(x):
    if isinstance(x, datetime):
        return x.time()
    if isinstance(x, time):
        return x

    x = clean(x).replace("上午", "AM ").replace("下午", "PM ")

    for fmt in ["%p %I:%M:%S", "%p %I:%M", "%H:%M:%S", "%H:%M"]:
        try:
            return datetime.strptime(x, fmt).time()
        except ValueError:
            pass

    raise ValueError(f"時間格式錯誤：{x}")


def parse_lab(x):
    x = clean(x)
    m = re.match(r"^(.*?)\s*-\s*(.*?)\s*\((.*?)\)\s*$", x)

    if m:
        area = clean(m.group(1))
        lab = clean(m.group(2))
        address = clean(m.group(3))
        return area, lab, address, f"{area} - {lab}"

    return "", x, "", x


def get(row, headers, name):
    name = name.strip()
    for key, col in headers.items():
        if key.strip() == name:
            return row[col - 1]
    raise KeyError(f"找不到欄位：{name}")


Path(OUTPUT_DIR).mkdir(exist_ok=True)

form_wb = load_workbook(FORM_FILE, data_only=True)
form_ws = form_wb.active

headers = {
    clean(cell.value): cell.column
    for cell in form_ws[1]
    if cell.value is not None
}

for row in form_ws.iter_rows(min_row=2, values_only=True):
    name = clean(get(row, headers, "姓名"))

    if not name:
        continue

    id_no = clean(get(row, headers, "身分證(護照)號"))
    birthday = parse_date(get(row, headers, "生日"))
    sample_date = parse_date(get(row, headers, "請填寫預約採檢日期"))
    sample_time = parse_time(get(row, headers, "請填寫預計預約採檢時間"))

    area, lab, address, display_lab = parse_lab(
        get(row, headers, "請選擇預約採檢單位")
    )

    wb = load_workbook(TEMPLATE_FILE)
    ws = wb[SHEET_NAME]

    ws[CELLS["姓名"]] = name
    ws[CELLS["身分證"]] = id_no
    ws[CELLS["生日"]] = birthday.strftime("%Y/%m/%d")
    ws[CELLS["檢驗所"]] = display_lab
    ws[CELLS["地址"]] = address

    ws[CELLS["年"]] = sample_date.year
    ws[CELLS["月"]] = sample_date.month
    ws[CELLS["日"]] = sample_date.day
    ws[CELLS["時"]] = sample_time.hour
    ws[CELLS["分"]] = f"{sample_time.minute:02d}"

    filename = f"{sample_date:%Y%m%d}_{safe_name(name)}_{safe_name(area + '-' + lab)}.xlsx"
    wb.save(Path(OUTPUT_DIR) / filename)

    print("已產生：", filename)

print("全部完成")